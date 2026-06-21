"""Import des vraies données des 3 fichiers Excel dans la base de données."""
import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

from datetime import datetime, date
import openpyxl
from app import app
from models import db, Operation, AuditLog

MIN_VALID_DATE = date(2000, 1, 1)

def parse_date(val):
    """Convertit une valeur Excel en date Python."""
    if val is None or val == '/' or val == '' or val == '-':
        return None
    parsed = None
    if isinstance(val, datetime):
        parsed = val.date()
    elif isinstance(val, date):
        parsed = val
    # Essayer de parser une string
    if parsed is None:
        try:
            parsed = datetime.strptime(str(val).strip(), '%Y-%m-%d').date()
        except:
            pass
    if parsed is None:
        try:
            parsed = datetime.strptime(str(val).strip(), '%d/%m/%Y').date()
        except:
            pass

    if parsed and parsed < MIN_VALID_DATE:
        return None
    return parsed

def parse_float(val):
    """Convertit une valeur en float."""
    if val is None or val == '' or val == '/' or val == '-':
        return None
    try:
        return float(val)
    except:
        return None

def clean_str(val):
    """Nettoie une string."""
    if val is None or val == '/' or val == '-':
        return None
    s = str(val).strip()
    return s if s else None

def normalize_statut(val):
    """Normalise le statut."""
    if val is None or val == '' or val == '/':
        return 'En attente'
    s = str(val).strip().lower()
    if s in ('ok', 'encaissé', 'encaisse'):
        return 'Encaissé'
    if 'attente' in s or 'à échéance' in s or 'echeance' in s:
        return 'En attente'
    if s in ('garantie',):
        return 'En attente'  # type_detail sera Garantie
    if 'rejet' in s:
        return 'Rejeté'
    if 'annul' in s:
        return 'Annulé'
    if 'transfert' in s:
        return 'Encaissé'
    if 'en cours' in s:
        return 'En cours'
    return 'En attente'

def normalize_societe(val):
    """Normalise le nom de société."""
    if val is None:
        return 'SRID'
    s = str(val).strip().lower()
    if 'genetic' in s:
        return 'Genetics'
    if 'srid' in s or 'ent' in s:
        return 'SRID'
    return 'SRID'

def normalize_type_detail(val):
    """Normalise le type détail (cn, cc, etc.)."""
    if val is None or val == '' or val == '/':
        return None
    s = str(val).strip().lower()
    if s == 'cn':
        return 'Courant'
    if s == 'cc':
        return 'Courant'
    if 'garantie' in s:
        return 'Garantie'
    return clean_str(val)

def import_cheques_file():
    """Import fichier 'Chèques ----.xlsx'"""
    filepath = os.path.join(os.path.dirname(__file__), 'Chèques  SABRINA.xlsx')
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        # Trouver la ligne d'en-tête (contient "Client" ou "Remettant")
        header_row = None
        for i, row in enumerate(rows):
            row_str = ' '.join([str(c).lower() for c in row if c])
            if 'client' in row_str or 'remettant' in row_str:
                header_row = i
                break
        
        if header_row is None:
            continue

        # Colonnes: N°, Date Reception, Remettant, Banque, Chèque N°, Client, Date Chèque, Report encaissement, Montant, Statut, Société, Entrée, Sortie, Date Sortie, Remarque
        for row in rows[header_row + 1:]:
            if not row or len(row) < 9:
                continue
            
            montant = parse_float(row[8] if len(row) > 8 else None)
            date_op = parse_date(row[6] if len(row) > 6 else None)  # Date du chèque
            client = clean_str(row[5] if len(row) > 5 else None)
            
            if not montant or not client:
                continue
            if not date_op:
                date_op = parse_date(row[1] if len(row) > 1 else None)  # Fallback: date reception
            if not date_op:
                continue

            statut_raw = clean_str(row[9] if len(row) > 9 else None)
            remarque_raw = clean_str(row[14] if len(row) > 14 else None)
            
            # Si le statut brut est dans la colonne remarque (fichier 1 a "Encaissé" en col 14)
            if remarque_raw and remarque_raw.lower() in ('encaissé', 'ok', 'rejeté', 'annulé'):
                if not statut_raw or statut_raw == '/':
                    statut_raw = remarque_raw
                    remarque_raw = None

            type_detail = None
            if 'garantie' in sheet_name.lower():
                type_detail = 'Garantie'
            
            # Vérifier si le statut brut contient "Garantie"
            if statut_raw and 'garantie' in statut_raw.lower():
                type_detail = 'Garantie'

            op = Operation(
                type_operation='Chèque',
                societe=normalize_societe(row[10] if len(row) > 10 else None),
                date_operation=date_op,
                date_reception=parse_date(row[1] if len(row) > 1 else None),
                date_encaissement=parse_date(row[7] if len(row) > 7 else None),
                date_sortie=parse_date(row[13] if len(row) > 13 else None),
                client=client,
                remettant=clean_str(row[2] if len(row) > 2 else None),
                montant=montant,
                banque=clean_str(row[3] if len(row) > 3 else None),
                numero_piece=clean_str(row[4] if len(row) > 4 else None),
                statut=normalize_statut(statut_raw),
                type_detail=type_detail,
                entree=clean_str(row[11] if len(row) > 11 else None),
                sortie=clean_str(row[12] if len(row) > 12 else None),
                remarque=remarque_raw,
                cree_par='Import Excel',
            )
            db.session.add(op)
            count += 1

    wb.close()
    return count

def import_ent_genetics_file():
    """Import fichier 'ENT GENETICS.xlsx'"""
    filepath = os.path.join(os.path.dirname(__file__), 'SRID GENETICS.xlsx')
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        # Trouver la ligne d'en-tête
        header_row = None
        for i, row in enumerate(rows):
            row_str = ' '.join([str(c).lower() for c in row if c])
            if 'client' in row_str or 'montant' in row_str:
                header_row = i
                break
        
        if header_row is None:
            continue

        # Lire les en-têtes pour mapping dynamique
        headers = [str(c).strip().lower() if c else '' for c in rows[header_row]]
        
        for row in rows[header_row + 1:]:
            if not row or len(row) < 5:
                continue
            
            # Mapping par position connue: '', Societé, Mode paiement, Date, Clients, Montant, Banque, Type/Statut...
            societe = clean_str(row[1] if len(row) > 1 else None)
            type_op = clean_str(row[2] if len(row) > 2 else None)
            date_op = parse_date(row[3] if len(row) > 3 else None)
            client = clean_str(row[4] if len(row) > 4 else None)
            montant = parse_float(row[5] if len(row) > 5 else None)
            banque = clean_str(row[6] if len(row) > 6 else None)
            
            if not montant or not client or not date_op:
                continue

            # Déterminer type_detail et statut selon les colonnes restantes
            type_detail = None
            statut_raw = None
            
            if 'type' in headers:
                idx = headers.index('type')
                type_detail = normalize_type_detail(row[idx] if len(row) > idx else None)
            
            # Statut: chercher dans les colonnes après banque
            for idx in range(7, min(len(row), 10)):
                val = clean_str(row[idx] if len(row) > idx else None)
                if val and val.lower() in ('ok', 'encaissé', 'rejeté', 'annulé', 'en cours', 'transfert'):
                    statut_raw = val
                    break

            # Normaliser le type d'opération
            if type_op:
                t = type_op.lower()
                if 'chèque' in t or 'cheque' in t:
                    type_op = 'Chèque'
                elif 'virement' in t:
                    type_op = 'Virement'
                elif 'versement' in t:
                    type_op = 'Versement'
            else:
                # Deviner depuis le nom de la feuille
                sn = sheet_name.lower()
                if 'virement' in sn or 'versement' in sn:
                    type_op = 'Virement'
                else:
                    type_op = 'Chèque'

            op = Operation(
                type_operation=type_op,
                societe=normalize_societe(societe),
                date_operation=date_op,
                client=client,
                montant=montant,
                banque=banque,
                statut=normalize_statut(statut_raw),
                type_detail=type_detail,
                cree_par='Import Excel',
            )
            db.session.add(op)
            count += 1

    wb.close()
    return count

def import_versement_cheques_ent():
    """Import fichier 'versement et cheques ENT serveur .xlsx'"""
    filepath = os.path.join(os.path.dirname(__file__), 'versement et cheques srid serveur .xlsx')
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        # Trouver la ligne d'en-tête
        header_row = None
        for i, row in enumerate(rows):
            row_str = ' '.join([str(c).lower() for c in row if c])
            if 'client' in row_str or 'montant' in row_str or 'nom' in row_str or 'libellé' in row_str:
                header_row = i
                break
        
        if header_row is None:
            continue

        headers = [str(c).strip().lower() if c else '' for c in rows[header_row]]
        
        # Déterminer le type d'opération depuis le nom de la feuille
        sn = sheet_name.lower()
        if 'virement' in sn and 'versement' in sn:
            default_type = None  # Mixte - devra lire la colonne type
        elif 'virement' in sn:
            default_type = 'Virement'
        elif 'versement' in sn:
            default_type = 'Versement'
        elif 'cheq' in sn or 'chèque' in sn:
            default_type = 'Chèque'
        else:
            # Feuilles année simple (2011, 2012...) = versements/chèques ENT
            default_type = 'Chèque'

        # Mapper les colonnes
        def find_col(names):
            for name in names:
                for idx, h in enumerate(headers):
                    if name in h:
                        return idx
            return None

        col_date = find_col(['date'])
        col_client = find_col(['client', 'nom', 'libellé'])
        col_montant = find_col(['montant'])
        col_banque = find_col(['banque', 'etabli', 'lieu'])
        col_piece = find_col(['n°', 'chéque', 'cheque', 'vers'])
        col_type = find_col(['type', 'mode'])
        col_statut = find_col(['statut'])
        col_famille = find_col(['famille'])
        col_remarque = find_col(['remarque'])

        for row in rows[header_row + 1:]:
            if not row or all(v is None or v == '' for v in row):
                continue
            
            date_op = parse_date(row[col_date]) if col_date is not None and len(row) > col_date else None
            client = clean_str(row[col_client]) if col_client is not None and len(row) > col_client else None
            montant = parse_float(row[col_montant]) if col_montant is not None and len(row) > col_montant else None
            
            if not montant or not client or not date_op:
                continue

            banque = clean_str(row[col_banque]) if col_banque is not None and len(row) > col_banque else None
            numero = clean_str(row[col_piece]) if col_piece is not None and len(row) > col_piece else None
            famille = clean_str(row[col_famille]) if col_famille is not None and len(row) > col_famille else None
            remarque = clean_str(row[col_remarque]) if col_remarque is not None and len(row) > col_remarque else None
            
            # Type
            type_op = default_type
            if col_type is not None and len(row) > col_type:
                t = clean_str(row[col_type])
                if t:
                    tl = t.lower()
                    if 'virement' in tl:
                        type_op = 'Virement'
                    elif 'versement' in tl:
                        type_op = 'Versement'
                    elif 'cheque' in tl or 'chèque' in tl:
                        type_op = 'Chèque'
                    else:
                        type_op = type_op or 'Chèque'
            if not type_op:
                type_op = 'Chèque'

            # Statut
            statut_raw = None
            if col_statut is not None and len(row) > col_statut:
                statut_raw = clean_str(row[col_statut])
            
            # Type détail (cn, cc)
            type_detail = None
            if col_type is not None and len(row) > col_type:
                td = clean_str(row[col_type])
                if td and td.lower() in ('cn', 'cc'):
                    type_detail = normalize_type_detail(td)

            # Numéro de pièce: si c'est "virement" c'est pas un n° pièce
            if numero and numero.lower() == 'virement':
                numero = None

            op = Operation(
                type_operation=type_op,
                societe=normalize_societe(famille) if famille else 'SRID',
                famille=famille,
                date_operation=date_op,
                client=client,
                montant=montant,
                banque=banque,
                numero_piece=numero,
                statut=normalize_statut(statut_raw),
                type_detail=type_detail,
                remarque=remarque,
                cree_par='Import Excel',
            )
            db.session.add(op)
            count += 1

    wb.close()
    return count


if __name__ == '__main__':
    with app.app_context():
        # Backup file before refresh
        db_path = os.path.join(os.path.dirname(__file__), 'database.db')
        if os.path.exists(db_path):
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = os.path.join(os.path.dirname(__file__), f'database_backup_{ts}.db')
            import shutil
            shutil.copy2(db_path, backup_path)
            print(f"Backup créé: {backup_path}")

        # Refresh operations and audit log from latest Excel sources.
        AuditLog.query.delete()
        Operation.query.delete()
        db.session.commit()

        print("Import des fichiers Excel...")
        
        c1 = import_cheques_file()
        print(f"  ✓ Chèques  SABRINA.xlsx : {c1} opérations")
        
        c2 = import_ent_genetics_file()
        print(f"  ✓ SRID GENETICS.xlsx : {c2} opérations")
        
        c3 = import_versement_cheques_ent()
        print(f"  ✓ versement et cheques srid serveur .xlsx : {c3} opérations")
        
        db.session.commit()
        
        total = Operation.query.count()
        print(f"\n{'='*50}")
        print(f"  TOTAL importé : {total} opérations")
        print(f"{'='*50}")
