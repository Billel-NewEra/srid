from flask import Flask, render_template, request, redirect, url_for, jsonify, session, flash, send_file, send_from_directory, make_response
from config import Config
from models import db, Operation, User, AuditLog, ClientLabel, RemettantLabel, CommandeLogistique, BonCommande, LigneCommande, Fournisseur, Product
from datetime import datetime, date, timedelta
from sqlalchemy import or_, func, extract, desc, case
from functools import wraps
import io
import re
import json
import os
from openpyxl import Workbook, load_workbook

REF_PER_PAGE = 25
LOG_PER_PAGE = 25
BON_PER_PAGE = 25
BON_STATUTS  = ['Brouillon', 'En attente', 'Approuvé', 'Envoyé', 'Reçu']


app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)

@app.template_filter('currency')
def currency_filter(value, decimals=2):
    """Formate un montant en format français : 1\u2009234\u2009567,89 (ou sans centimes si decimals=0)"""
    try:
        formatted = "{:,.{d}f}".format(float(value), d=decimals)
        # virgule -> espace fine (milliers), point -> virgule (décimale)
        formatted = formatted.replace(",", "\u2009").replace(".", ",")
        return formatted
    except (TypeError, ValueError):
        return value


ROLE_ALIASES = {
    'boss': 'admin',
    'admin': 'admin',
    'saisisseur': 'saisie',
    'saisie': 'saisie',
    'consultation': 'consultation',
    'consultant': 'consultation',
    'viewer': 'consultation',
}

ROLE_LABELS = {
    'admin': 'Admin',
    'saisie': 'Saisie',
    'consultation': 'Consultation',
}

CHECK_TYPE_CHOICES = ['Garantie', 'À encaisser', 'À échéance']
STATUS_CHOICES = ['Encaissé', 'Rejeté', 'Échéance', 'En cours', 'Arrive à échéance', 'Échu']


# --- Décorateurs d'authentification ---


def _normalize_role(role):
    return ROLE_ALIASES.get((role or '').strip().lower(), 'consultation')


def _current_role():
    return _normalize_role(session.get('user_role', ''))


def _forbidden_response():
    if request.path.startswith('/api/') or request.headers.get('HX-Request'):
        return ('Accès refusé.', 403)
    flash('Accès refusé pour ce rôle.', 'error')
    return redirect(url_for('dashboard'))

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def role_required(*allowed_roles):
    allowed = {_normalize_role(r) for r in allowed_roles}

    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            if _current_role() not in allowed:
                return _forbidden_response()
            return f(*args, **kwargs)

        return decorated_function

    return decorator


# --- Contexte global templates ---

@app.context_processor
def inject_globals():
    from datetime import datetime
    role_key = _current_role()
    return {
        'current_user_id': session.get('user_id'),
        'current_user': session.get('user_nom', ''),
        'current_role': ROLE_LABELS.get(role_key, 'Consultation'),
        'current_role_key': role_key,
        'can_write': role_key in {'admin', 'saisie'},
        'can_delete': role_key == 'admin',
        'is_admin': role_key == 'admin',
        'can_manage_users': role_key == 'admin',
        'is_logged_in': 'user_id' in session,
        'now': datetime.now,
    }


def _normalize_bank_name(raw):
    """Normalise les variantes de banques vers un nom canonique."""
    if raw is None:
        return None
    source = str(raw).strip()
    if not source:
        return None

    cleaned = " ".join(source.split()).upper()
    compact = re.sub(r'[^A-Z0-9]', '', cleaned)
    compact = re.sub(r'\d+$', '', compact)
    if not compact:
        return None

    if compact.startswith('BADR') or compact.startswith('BAXDR'):
        return 'BADR'
    if compact.startswith('HOUSING') or compact in {'HB', 'HH', 'HBTF', 'HTBF'}:
        return 'HOUSING BANK'
    if compact.startswith('BNA'):
        return 'BNA'
    if compact.startswith('BDL'):
        return 'BDL'
    if compact.startswith('CPA'):
        return 'CPA'
    if compact.startswith('BEA'):
        return 'BEA'
    if compact.startswith('AGB') or compact.startswith('GBA'):
        return 'AGB'
    if compact.startswith('SGA') or compact == 'SG':
        return 'SGA'
    if compact.startswith('ALSALAMBANK') or compact.startswith('SALAMBANK'):
        return 'AL SALAM BANK'
    if compact.startswith('ALBARAKA') or compact.startswith('ELBARAKA'):
        return 'AL BARAKA'
    if compact.startswith('TRUST'):
        return 'TRUST BANK'
    if compact.startswith('FRANSABANK') or compact == 'FB':
        return 'FRANSABANK'
    if compact.startswith('ARABBANK'):
        return 'ARAB BANK'
    if compact.startswith('CNEP'):
        return 'CNEP'
    if compact.startswith('CCP'):
        return 'CCP'
    if compact.startswith('CITIBANK'):
        return 'CITIBANK'
    if compact.startswith('BNH'):
        return 'BNH'

    return cleaned


def _get_bank_suggestions():
    """Construit une liste unique de banques normalisées pour les formulaires."""
    rows = db.session.query(Operation.banque).filter(Operation.banque.isnot(None)).all()
    values = {_normalize_bank_name(v) for (v,) in rows}
    values.discard(None)

    defaults = {
        'BADR', 'HOUSING BANK', 'BNA', 'BDL', 'CPA', 'BEA',
        'AGB', 'SGA', 'AL BARAKA', 'AL SALAM BANK', 'TRUST BANK',
        'FRANSABANK', 'ARAB BANK', 'CNEP'
    }
    values.update(defaults)
    return sorted(values)


def _get_client_suggestions():
    # Depuis les opérations existantes
    rows = db.session.query(Operation.client).filter(Operation.client.isnot(None), Operation.client != '').all()
    values = {str(v).strip() for (v,) in rows if v}
    # Depuis le référentiel dédié
    labels = ClientLabel.query.filter_by(actif=True).all()
    values.update(l.nom.strip() for l in labels)
    return sorted(values)


def _get_remettant_suggestions():
    # Depuis les opérations existantes
    rows = db.session.query(Operation.remettant).filter(Operation.remettant.isnot(None), Operation.remettant != '').all()
    values = {str(v).strip() for (v,) in rows if v}
    # Depuis le référentiel dédié
    labels = RemettantLabel.query.filter_by(actif=True).all()
    values.update(l.nom.strip() for l in labels)
    return sorted(values)


def _normalize_type_operation(raw_type):
    s = (raw_type or '').strip().lower()
    if s in {'chèque', 'cheque'}:
        return 'Chèque'
    if s == 'virement':
        return 'Virement'
    if s == 'versement':
        return 'Versement'
    if s == 'transfer':
        return 'Transfer'
    if s == 'autre':
        return 'Autre'
    return 'Autre'


def _compute_statut(type_operation, type_cheque):
    if type_operation != 'Chèque':
        return 'Encaissé'
    if type_cheque == 'À échéance':
        return 'Échéance'
    return 'En cours'


def _normalize_legacy_statut(op):
    raw = (op.statut or '').strip()
    if raw in STATUS_CHOICES:
        return raw

    lowered = raw.lower()
    if 'annul' in lowered:
        return 'Rejeté'
    if 'rejet' in lowered:
        return 'Rejeté'
    if 'ech' in lowered:
        return 'Échéance'

    type_cheque = op.type_detail if op.type_operation == 'Chèque' else None
    return _compute_statut(op.type_operation, type_cheque)


def _auto_update_echeance_statuts():
    """Passe automatiquement les statuts des chèques à échéance en fonction de la date."""
    today = date.today()
    alert_date = today + timedelta(days=7)
    # Échu : date passée uniquement pour les statuts de workflow échéance.
    # Ne pas écraser un statut manuel (ex: Encaissé/Rejeté) choisi par l'admin.
    db.session.query(Operation).filter(
        Operation.type_operation == 'Chèque',
        Operation.type_detail == 'À échéance',
        Operation.date_encaissement < today,
        Operation.statut.in_(['Échéance', 'Arrive à échéance']),
    ).update({'statut': 'Échu'}, synchronize_session=False)
    # Arrive à échéance : dans les 7 prochains jours, statut encore Échéance
    db.session.query(Operation).filter(
        Operation.type_operation == 'Chèque',
        Operation.type_detail == 'À échéance',
        Operation.date_encaissement >= today,
        Operation.date_encaissement <= alert_date,
        Operation.statut == 'Échéance',
    ).update({'statut': 'Arrive à échéance'}, synchronize_session=False)
    db.session.commit()


def _get_echeance_notifications(limit=8):
    """Retourne les alertes d'échéance globales pour affichage sans filtre."""
    today = date.today()
    alert_date = today + timedelta(days=7)
    workflow_statuses = ['Échéance', 'Arrive à échéance', 'Échu']
    base_query = Operation.query.filter(
        Operation.type_operation == 'Chèque',
        Operation.type_detail == 'À échéance',
        Operation.date_encaissement.isnot(None),
        Operation.statut.in_(workflow_statuses),
    )

    overdue_count = base_query.filter(Operation.date_encaissement < today).count()
    upcoming_count = base_query.filter(
        Operation.date_encaissement >= today,
        Operation.date_encaissement <= alert_date,
    ).count()

    critical_operations = base_query.filter(
        or_(
            Operation.date_encaissement < today,
            (
                (Operation.date_encaissement >= today) &
                (Operation.date_encaissement <= alert_date)
            ),
        )
    ).order_by(Operation.date_encaissement.asc()).limit(limit).all()

    return {
        'today': today,
        'overdue_count': overdue_count,
        'upcoming_count': upcoming_count,
        'total_alerts': overdue_count + upcoming_count,
        'critical_operations': critical_operations,
    }


def _get_recent_rejections(limit=8):
    """Retourne les rejets récents (dernières 24h) depuis l'audit log."""
    now = datetime.utcnow()
    yesterday = now - timedelta(hours=24)
    
    # Chercher les audits où les details contiennent "Rejeté" et la date_action est récente
    rejections = AuditLog.query.filter(
        AuditLog.date_action >= yesterday,
        AuditLog.details.ilike('%Rejeté%'),
    ).order_by(desc(AuditLog.date_action)).limit(limit).all()
    
    # Grouper par operation_id et récupérer les infos de l'opération
    rejection_ops = []
    seen_ops = set()
    for audit in rejections:
        if audit.operation_id not in seen_ops:
            op = Operation.query.get(audit.operation_id)
            if op and op.statut == 'Rejeté':
                rejection_ops.append({
                    'operation': op,
                    'rejected_at': audit.date_action,
                    'rejected_by': audit.utilisateur,
                })
                seen_ops.add(audit.operation_id)
    
    return {
        'rejections': rejection_ops,
        'total': len(rejection_ops),
    }


def _admin_users_count():
    return User.query.filter_by(role='admin').count()


# --- Routes Auth ---

@app.route('/sw.js')
def service_worker():
    return send_from_directory('static', 'sw.js', mimetype='application/javascript')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['user_nom'] = user.nom_complet or user.username
            session['user_role'] = _normalize_role(user.role)
            flash(f'Bienvenue, {user.nom_complet or user.username} !', 'success')
            return redirect(url_for('dashboard'))
        flash('Identifiants incorrects.', 'error')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/utilisateurs', methods=['GET', 'POST'])
@role_required('admin')
def manage_users():
    role_choices = ['admin', 'saisie', 'consultation']

    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        nom_complet = request.form.get('nom_complet', '').strip()
        password = request.form.get('password', '')
        role = _normalize_role(request.form.get('role', 'consultation'))

        if not username or len(username) < 3:
            flash('Nom utilisateur invalide (min 3 caractères).', 'error')
            return redirect(url_for('manage_users'))
        if role not in role_choices:
            flash('Rôle invalide.', 'error')
            return redirect(url_for('manage_users'))
        if len(password) < 6:
            flash('Mot de passe trop court (min 6 caractères).', 'error')
            return redirect(url_for('manage_users'))
        if User.query.filter_by(username=username).first():
            flash('Ce nom utilisateur existe déjà.', 'error')
            return redirect(url_for('manage_users'))

        user = User(
            username=username,
            nom_complet=nom_complet or username,
            role=role,
        )
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash(f'Utilisateur {username} créé avec rôle {ROLE_LABELS.get(role, role)}.', 'success')
        return redirect(url_for('manage_users'))

    users = User.query.order_by(User.username.asc()).all()
    return render_template('users.html', users=users, role_labels=ROLE_LABELS)


@app.route('/utilisateurs/<int:user_id>/role', methods=['POST'])
@role_required('admin')
def update_user_role(user_id):
    user = User.query.get_or_404(user_id)
    new_role = _normalize_role(request.form.get('role', 'consultation'))
    if new_role not in {'admin', 'saisie', 'consultation'}:
        flash('Rôle invalide.', 'error')
        return redirect(url_for('manage_users'))

    if user.id == session.get('user_id'):
        flash('Vous ne pouvez pas modifier votre propre rôle.', 'error')
        return redirect(url_for('manage_users'))

    if user.role == 'admin' and new_role != 'admin' and _admin_users_count() <= 1:
        flash('Impossible de retirer le dernier administrateur.', 'error')
        return redirect(url_for('manage_users'))

    user.role = new_role
    db.session.commit()
    flash(
        f'Rôle mis à jour pour {user.username}. Le nouvel accès sera effectif après reconnexion de cet utilisateur.',
        'success'
    )
    return redirect(url_for('manage_users'))


@app.route('/utilisateurs/<int:user_id>/password', methods=['POST'])
@role_required('admin')
def update_user_password(user_id):
    user = User.query.get_or_404(user_id)
    new_password = request.form.get('password', '')
    if len(new_password) < 6:
        flash('Mot de passe trop court (min 6 caractères).', 'error')
        return redirect(url_for('manage_users'))

    user.set_password(new_password)
    db.session.commit()
    flash(
        f'Mot de passe mis à jour pour {user.username}. Le changement sera pris en compte après reconnexion.',
        'success'
    )
    return redirect(url_for('manage_users'))


@app.route('/utilisateurs/<int:user_id>/delete', methods=['POST'])
@role_required('admin')
def delete_user(user_id):
    user = User.query.get_or_404(user_id)

    if user.id == session.get('user_id'):
        flash('Vous ne pouvez pas supprimer votre propre compte.', 'error')
        return redirect(url_for('manage_users'))

    if user.role == 'admin' and _admin_users_count() <= 1:
        flash('Impossible de supprimer le dernier administrateur.', 'error')
        return redirect(url_for('manage_users'))

    username = user.username
    db.session.delete(user)
    db.session.commit()
    flash(f'Utilisateur {username} supprimé.', 'success')
    return redirect(url_for('manage_users'))


# --- Dashboard ---

@app.route('/')
@login_required
def dashboard():
    current_year = date.today().year
    previous_year = current_year - 1
    today = date.today()
    status_labels = STATUS_CHOICES

    # Années disponibles
    years_raw = db.session.query(
        extract('year', Operation.date_operation)
    ).distinct().order_by(extract('year', Operation.date_operation).desc()).all()
    available_years = [int(y[0]) for y in years_raw if y[0]]

    # --- KPIs globaux (1 query) ---
    kpi_raw = db.session.query(
        Operation.societe,
        func.coalesce(func.sum(Operation.montant), 0)
    ).group_by(Operation.societe).all()
    kpi_lookup = {s: float(m) for s, m in kpi_raw}
    total_montant = sum(kpi_lookup.values())
    montant_srid = kpi_lookup.get('SRID', 0.0)
    montant_genetics = kpi_lookup.get('Genetics', 0.0)

    # --- Statuts avec montants par société (1 query) ---
    statuts_raw = db.session.query(
        Operation.statut,
        Operation.societe,
        func.count(Operation.id),
        func.coalesce(func.sum(Operation.montant), 0)
    ).group_by(Operation.statut, Operation.societe).all()
    statuts_info = {s: {'count': 0, 'montant': 0.0, 'srid_count': 0, 'srid_montant': 0.0,
                        'genetics_count': 0, 'genetics_montant': 0.0} for s in status_labels}
    for statut, societe, count, montant in statuts_raw:
        if statut in statuts_info:
            statuts_info[statut]['count'] += count
            statuts_info[statut]['montant'] += float(montant)
            if societe == 'SRID':
                statuts_info[statut]['srid_count'] += count
                statuts_info[statut]['srid_montant'] += float(montant)
            elif societe == 'Genetics':
                statuts_info[statut]['genetics_count'] += count
                statuts_info[statut]['genetics_montant'] += float(montant)

    # --- Types avec montants par société (1 query) ---
    types_raw = db.session.query(
        Operation.type_operation,
        Operation.societe,
        func.count(Operation.id),
        func.coalesce(func.sum(Operation.montant), 0)
    ).group_by(Operation.type_operation, Operation.societe).all()
    type_list = ['Chèque', 'Virement', 'Versement', 'Transfer', 'Autre']
    types_info = {t: {'total': {'count': 0, 'montant': 0.0},
                      'SRID': {'count': 0, 'montant': 0.0},
                      'Genetics': {'count': 0, 'montant': 0.0}} for t in type_list}
    for type_op, societe, count, montant in types_raw:
        if type_op in types_info:
            types_info[type_op]['total']['count'] += count
            types_info[type_op]['total']['montant'] += float(montant)
            if societe in ('SRID', 'Genetics'):
                types_info[type_op][societe]['count'] += count
                types_info[type_op][societe]['montant'] += float(montant)

    # --- Top 5 clients (année en cours) ---
    top_clients = db.session.query(
        Operation.client,
        func.sum(Operation.montant).label('total')
    ).filter(
        extract('year', Operation.date_operation) == current_year,
        Operation.client.isnot(None),
        Operation.client != '',
        ~func.lower(Operation.client).like('%srid%'),
        ~func.lower(Operation.client).like('%genetics%')
    ).group_by(Operation.client).order_by(desc('total')).limit(5).all()

    # --- Dernières opérations ---
    dernieres = Operation.query.order_by(Operation.date_operation.desc()).limit(10).all()

    # --- Données mensuelles année courante + par société (1 query) ---
    monthly_curr_raw = db.session.query(
        extract('month', Operation.date_operation).label('month'),
        Operation.societe,
        func.coalesce(func.sum(Operation.montant), 0)
    ).filter(
        extract('year', Operation.date_operation) == current_year
    ).group_by('month', Operation.societe).all()
    monthly_curr_lookup = {}
    for month, societe, montant in monthly_curr_raw:
        m = int(month)
        if m not in monthly_curr_lookup:
            monthly_curr_lookup[m] = {'total': 0.0, 'SRID': 0.0, 'Genetics': 0.0}
        monthly_curr_lookup[m]['total'] += float(montant)
        if societe in ('SRID', 'Genetics'):
            monthly_curr_lookup[m][societe] += float(montant)
    monthly_data = [monthly_curr_lookup.get(m, {}).get('total', 0.0) for m in range(1, 13)]
    monthly_srid = [monthly_curr_lookup.get(m, {}).get('SRID', 0.0) for m in range(1, 13)]
    monthly_genetics = [monthly_curr_lookup.get(m, {}).get('Genetics', 0.0) for m in range(1, 13)]

    # --- Données mensuelles année précédente (1 query) ---
    monthly_prev_raw = db.session.query(
        extract('month', Operation.date_operation).label('month'),
        func.coalesce(func.sum(Operation.montant), 0)
    ).filter(
        extract('year', Operation.date_operation) == previous_year
    ).group_by('month').all()
    monthly_prev_lookup = {int(m): float(mt) for m, mt in monthly_prev_raw}
    monthly_data_prev = [monthly_prev_lookup.get(m, 0.0) for m in range(1, 13)]

    # --- Statuts par mois année courante (1 query) ---
    monthly_statuts_raw = db.session.query(
        extract('month', Operation.date_operation).label('month'),
        Operation.statut,
        Operation.societe,
        func.coalesce(func.sum(Operation.montant), 0)
    ).filter(
        extract('year', Operation.date_operation) == current_year
    ).group_by('month', Operation.statut, Operation.societe).all()
    monthly_statuts_views = {
        'total': {s: [0.0] * 12 for s in status_labels},
        'srid':  {s: [0.0] * 12 for s in status_labels},
        'genetics': {s: [0.0] * 12 for s in status_labels},
    }
    for month, statut, societe, montant in monthly_statuts_raw:
        m = int(month) - 1
        if statut in status_labels:
            monthly_statuts_views['total'][statut][m] += float(montant)
            if societe == 'SRID':
                monthly_statuts_views['srid'][statut][m] += float(montant)
            elif societe == 'Genetics':
                monthly_statuts_views['genetics'][statut][m] += float(montant)
    monthly_statuts = monthly_statuts_views['total']

    # --- Activité 30 jours (1 query) ---
    thirty_days_ago = today - timedelta(days=29)
    daily_raw = db.session.query(
        func.date(Operation.date_operation).label('day'),
        Operation.societe,
        func.coalesce(func.sum(Operation.montant), 0)
    ).filter(
        func.date(Operation.date_operation) >= thirty_days_ago,
        func.date(Operation.date_operation) <= today
    ).group_by('day', Operation.societe).all()
    daily_lookup = {}
    for day, societe, montant in daily_raw:
        d_str = str(day)
        if d_str not in daily_lookup:
            daily_lookup[d_str] = {'total': 0.0, 'SRID': 0.0, 'Genetics': 0.0}
        daily_lookup[d_str]['total'] += float(montant)
        if societe in ('SRID', 'Genetics'):
            daily_lookup[d_str][societe] += float(montant)
    daily_labels, daily_total, daily_srid, daily_genetics = [], [], [], []
    for i in range(29, -1, -1):
        d = today - timedelta(days=i)
        entry = daily_lookup.get(d.isoformat(), {'total': 0.0, 'SRID': 0.0, 'Genetics': 0.0})
        daily_labels.append(d.strftime('%d/%m'))
        daily_total.append(entry['total'])
        daily_srid.append(entry['SRID'])
        daily_genetics.append(entry['Genetics'])

    return render_template('dashboard.html',
                           total_montant=float(total_montant),
                           montant_srid=float(montant_srid), montant_genetics=float(montant_genetics),
                           statuts_info=statuts_info, types_info=types_info,
                           top_clients=top_clients, dernieres=dernieres,
                           monthly_data=monthly_data, monthly_data_prev=monthly_data_prev,
                           monthly_srid=monthly_srid, monthly_genetics=monthly_genetics,
                           monthly_statuts=monthly_statuts,
                           monthly_statuts_views=monthly_statuts_views,
                           daily_labels=daily_labels, daily_total=daily_total,
                           daily_srid=daily_srid, daily_genetics=daily_genetics,
                           today=date.today().strftime('%d/%m/%Y'),
                           year=current_year, previous_year=previous_year,
                           available_years=available_years)


# --- API Dashboard (filtres HTMX/JS) ---

@app.route('/api/dashboard/monthly')
@login_required
def api_dashboard_monthly():
    year = request.args.get('year', date.today().year, type=int)
    prev_year = year - 1
    monthly_data = []
    monthly_data_prev = []
    for month in range(1, 13):
        total = db.session.query(func.sum(Operation.montant)).filter(
            extract('year', Operation.date_operation) == year,
            extract('month', Operation.date_operation) == month
        ).scalar() or 0
        total_prev = db.session.query(func.sum(Operation.montant)).filter(
            extract('year', Operation.date_operation) == prev_year,
            extract('month', Operation.date_operation) == month
        ).scalar() or 0
        monthly_data.append(float(total))
        monthly_data_prev.append(float(total_prev))
    return jsonify({'year': year, 'prev_year': prev_year, 'data': monthly_data, 'data_prev': monthly_data_prev})


@app.route('/api/dashboard/societes')
@login_required
def api_dashboard_societes():
    year = request.args.get('year', date.today().year, type=int)
    monthly_srid = []
    monthly_genetics = []
    for month in range(1, 13):
        t_srid = db.session.query(func.sum(Operation.montant)).filter(
            extract('year', Operation.date_operation) == year,
            extract('month', Operation.date_operation) == month,
            Operation.societe == 'SRID'
        ).scalar() or 0
        t_genetics = db.session.query(func.sum(Operation.montant)).filter(
            extract('year', Operation.date_operation) == year,
            extract('month', Operation.date_operation) == month,
            Operation.societe == 'Genetics'
        ).scalar() or 0
        monthly_srid.append(float(t_srid))
        monthly_genetics.append(float(t_genetics))
    return jsonify({'year': year, 'srid': monthly_srid, 'genetics': monthly_genetics})


@app.route('/api/dashboard/statuts-monthly')
@login_required
def api_dashboard_statuts_monthly():
    year = request.args.get('year', date.today().year, type=int)
    status_labels = STATUS_CHOICES
    monthly_statuts_views = {
        'total': {s: [] for s in status_labels},
        'srid': {s: [] for s in status_labels},
        'genetics': {s: [] for s in status_labels},
    }
    for month in range(1, 13):
        for statut in status_labels:
            total_mt = db.session.query(func.sum(Operation.montant)).filter(
                extract('year', Operation.date_operation) == year,
                extract('month', Operation.date_operation) == month,
                Operation.statut == statut
            ).scalar() or 0
            srid_mt = db.session.query(func.sum(Operation.montant)).filter(
                extract('year', Operation.date_operation) == year,
                extract('month', Operation.date_operation) == month,
                Operation.statut == statut,
                Operation.societe == 'SRID'
            ).scalar() or 0
            genetics_mt = db.session.query(func.sum(Operation.montant)).filter(
                extract('year', Operation.date_operation) == year,
                extract('month', Operation.date_operation) == month,
                Operation.statut == statut,
                Operation.societe == 'Genetics'
            ).scalar() or 0
            monthly_statuts_views['total'][statut].append(float(total_mt))
            monthly_statuts_views['srid'][statut].append(float(srid_mt))
            monthly_statuts_views['genetics'][statut].append(float(genetics_mt))
    return jsonify({'year': year, 'data': monthly_statuts_views['total'], 'views': monthly_statuts_views})


@app.route('/api/dashboard/top-clients')
@login_required
def api_dashboard_top_clients():
    year = request.args.get('year', date.today().year, type=int)
    month = request.args.get('month', 0, type=int)
    filters = [
        extract('year', Operation.date_operation) == year,
        Operation.client.isnot(None),
        Operation.client != '',
        ~func.lower(Operation.client).like('%srid%'),
        ~func.lower(Operation.client).like('%genetics%')
    ]
    if month > 0:
        filters.append(extract('month', Operation.date_operation) == month)
    top_clients = db.session.query(
        Operation.client,
        func.sum(Operation.montant).label('total')
    ).filter(*filters).group_by(Operation.client).order_by(desc('total')).limit(5).all()
    max_total = float(top_clients[0][1]) if top_clients else 1
    html = ''
    for i, (client_name, total) in enumerate(top_clients, 1):
        pct = float(total) / max_total * 100
        html += f'''<div class="flex items-center gap-3">
            <div class="badge badge-sm badge-primary font-bold w-6 h-6">{i}</div>
            <div class="flex-1">
                <div class="flex justify-between items-center mb-1">
                    <span class="text-sm font-medium truncate max-w-[180px]">{client_name or 'N/A'}</span>
                    <span class="text-sm font-bold">{total:,.0f} DA</span>
                </div>
                <progress class="progress progress-primary w-full h-2" value="{pct}" max="100"></progress>
            </div>
        </div>'''
    if not top_clients:
        html = '<p class="text-center opacity-50 py-4">Aucune donnée</p>'
    return html


@app.route('/api/dashboard/types')
@login_required
def api_dashboard_types():
    societe = request.args.get('societe', 'all')
    result = {}
    for type_op in ['Chèque', 'Virement', 'Versement', 'Transfer', 'Autre']:
        q = db.session.query(func.sum(Operation.montant)).filter_by(type_operation=type_op)
        if societe != 'all':
            q = q.filter_by(societe=societe)
        result[type_op] = float(q.scalar() or 0)
    return jsonify(result)


# --- Saisie ---

@app.route('/saisie', methods=['GET', 'POST'])
@role_required('admin', 'saisie')
def saisie():
    if request.method == 'POST':
        type_operation = _normalize_type_operation(request.form.get('type_operation'))
        type_cheque = (request.form.get('type_cheque') or request.form.get('type_detail')) if type_operation == 'Chèque' else None
        if type_cheque not in CHECK_TYPE_CHOICES:
            type_cheque = None

        date_operation = None
        date_reception = None
        date_sortie = None
        date_echeance = None
        if type_operation == 'Chèque':
            date_reception = _parse_date(request.form.get('date_reception'))
            date_sortie = _parse_date(request.form.get('date_sortie'))
            if type_cheque == 'À échéance':
                date_echeance = _parse_date(request.form.get('date_echeance') or request.form.get('date_encaissement'))
            date_operation = date_sortie
        else:
            date_operation = _parse_date(request.form.get('date_operation'))

        statut_auto = _compute_statut(type_operation, type_cheque)

        op = Operation(
            type_operation=type_operation,
            societe=request.form.get('societe'),
            famille=None,
            date_operation=date_operation,
            date_reception=date_reception,
            date_encaissement=date_echeance,
            date_sortie=date_sortie,
            client=request.form.get('client'),
            remettant=request.form.get('remettant_commercial') or request.form.get('remettant') or None,
            montant=float(request.form.get('montant', 0)),
            banque=_normalize_bank_name(request.form.get('banque')),
            numero_piece=request.form.get('numero_piece') or None,
            statut=statut_auto,
            type_detail=type_cheque,
            entree=request.form.get('entree') or None,
            sortie=request.form.get('sortie') or None,
            remarque=request.form.get('remarque') or None,
            cree_par=session.get('user_nom', ''),
        )
        db.session.add(op)
        db.session.commit()
        _log_audit(op.id, 'création', f"{op.type_operation} - {op.client} - {op.montant}")

        if request.headers.get('HX-Request'):
            return render_template('partials/success_message.html', operation=op)
        flash('Opération enregistrée !', 'success')
        return redirect(url_for('saisie'))

    return render_template(
        'saisie.html',
        bank_options=_get_bank_suggestions(),
        client_options=_get_client_suggestions(),
        remettant_options=_get_remettant_suggestions(),
        check_type_options=CHECK_TYPE_CHOICES,
    )


# --- Modification ---

@app.route('/edit/<int:op_id>', methods=['GET', 'POST'])
@role_required('admin', 'saisie')
def edit_operation(op_id):
    op = Operation.query.get_or_404(op_id)
    if request.method == 'POST':
        type_operation = _normalize_type_operation(request.form.get('type_operation'))
        type_cheque = (request.form.get('type_cheque') or request.form.get('type_detail')) if type_operation == 'Chèque' else None
        if type_cheque not in CHECK_TYPE_CHOICES:
            type_cheque = None

        date_operation = None
        date_reception = None
        date_sortie = None
        date_echeance = None
        if type_operation == 'Chèque':
            date_reception = _parse_date(request.form.get('date_reception'))
            date_sortie = _parse_date(request.form.get('date_sortie'))
            if type_cheque == 'À échéance':
                date_echeance = _parse_date(request.form.get('date_echeance') or request.form.get('date_encaissement'))
            date_operation = date_sortie
        else:
            date_operation = _parse_date(request.form.get('date_operation'))

        op.type_operation = type_operation
        op.societe = request.form.get('societe')
        op.famille = None
        op.date_operation = date_operation
        op.date_reception = date_reception
        op.date_encaissement = date_echeance
        op.date_sortie = date_sortie
        op.client = request.form.get('client')
        op.remettant = request.form.get('remettant_commercial') or request.form.get('remettant') or None
        op.montant = float(request.form.get('montant', 0))
        op.banque = _normalize_bank_name(request.form.get('banque'))
        op.numero_piece = request.form.get('numero_piece') or None
        op.statut = _compute_statut(type_operation, type_cheque)
        op.type_detail = type_cheque
        op.entree = request.form.get('entree') or None
        op.sortie = request.form.get('sortie') or None
        op.remarque = request.form.get('remarque') or None
        op.date_modification = datetime.utcnow()
        db.session.commit()
        _log_audit(op.id, 'modification', f"Modifié par {session.get('user_nom', '')}")

        if request.headers.get('HX-Request'):
            return render_template('partials/success_message.html', operation=op, action='modifiée')
        flash('Opération modifiée !', 'success')
        return redirect(url_for('consultation'))

    if request.headers.get('HX-Request'):
        return render_template(
            'partials/edit_form.html',
            operation=op,
            bank_options=_get_bank_suggestions(),
            client_options=_get_client_suggestions(),
            remettant_options=_get_remettant_suggestions(),
            check_type_options=CHECK_TYPE_CHOICES,
        )
    return render_template(
        'edit.html',
        operation=op,
        bank_options=_get_bank_suggestions(),
        client_options=_get_client_suggestions(),
        remettant_options=_get_remettant_suggestions(),
        check_type_options=CHECK_TYPE_CHOICES,
    )


# ─── RÉFÉRENTIELS ────────────────────────────────────────────────────────────

def _ref_clients_ctx(is_admin, page=1, search=''):
    q = ClientLabel.query.order_by(ClientLabel.nom)
    if search:
        q = q.filter(ClientLabel.nom.ilike(f'%{search}%'))
    total = q.count()
    clients = q.offset((page - 1) * REF_PER_PAGE).limit(REF_PER_PAGE).all()
    return dict(clients=clients, is_admin=is_admin, page=page, search=search,
                total=total, total_pages=max(1, (total + REF_PER_PAGE - 1) // REF_PER_PAGE))


def _ref_remettants_ctx(is_admin, page=1, search=''):
    q = RemettantLabel.query.order_by(RemettantLabel.nom)
    if search:
        q = q.filter(RemettantLabel.nom.ilike(f'%{search}%'))
    total = q.count()
    remettants = q.offset((page - 1) * REF_PER_PAGE).limit(REF_PER_PAGE).all()
    return dict(remettants=remettants, is_admin=is_admin, page=page, search=search,
                total=total, total_pages=max(1, (total + REF_PER_PAGE - 1) // REF_PER_PAGE))


@app.route('/referentiels')
@role_required('admin', 'saisie')
def referentiels():
    return render_template('referentiels.html', is_admin=_current_role() == 'admin')


@app.route('/api/referentiels/clients/list')
@role_required('admin', 'saisie')
def api_ref_clients_list():
    page   = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    return render_template('partials/ref_clients_list.html',
                           **_ref_clients_ctx(_current_role() == 'admin', page, search))


@app.route('/api/referentiels/remettants/list')
@role_required('admin', 'saisie')
def api_ref_remettants_list():
    page   = request.args.get('page', 1, type=int)
    search = request.args.get('search', '').strip()
    return render_template('partials/ref_remettants_list.html',
                           **_ref_remettants_ctx(_current_role() == 'admin', page, search))


@app.route('/api/referentiels/clients/add', methods=['POST'])
@role_required('admin', 'saisie')
def api_ref_client_add():
    nom = (request.form.get('nom') or '').strip()
    if not nom:
        return '<p class="text-error text-sm">Nom requis.</p>', 400
    if ClientLabel.query.filter(db.func.lower(ClientLabel.nom) == nom.lower()).first():
        return '<p class="text-error text-sm">Ce client existe déjà.</p>', 400
    db.session.add(ClientLabel(nom=nom))
    db.session.commit()
    return render_template('partials/ref_clients_list.html',
                           **_ref_clients_ctx(_current_role() == 'admin'))


@app.route('/api/referentiels/clients/<int:item_id>/toggle', methods=['POST'])
@role_required('admin', 'saisie')
def api_ref_client_toggle(item_id):
    item = ClientLabel.query.get_or_404(item_id)
    item.actif = not item.actif
    db.session.commit()
    return render_template('partials/ref_clients_list.html',
                           **_ref_clients_ctx(_current_role() == 'admin'))


@app.route('/api/referentiels/clients/<int:item_id>/delete', methods=['DELETE', 'POST'])
@role_required('admin')
def api_ref_client_delete(item_id):
    item = ClientLabel.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    return render_template('partials/ref_clients_list.html',
                           **_ref_clients_ctx(_current_role() == 'admin'))


@app.route('/api/referentiels/remettants/add', methods=['POST'])
@role_required('admin', 'saisie')
def api_ref_remettant_add():
    nom = (request.form.get('nom') or '').strip()
    if not nom:
        return '<p class="text-error text-sm">Nom requis.</p>', 400
    if RemettantLabel.query.filter(db.func.lower(RemettantLabel.nom) == nom.lower()).first():
        return '<p class="text-error text-sm">Ce remettant existe déjà.</p>', 400
    db.session.add(RemettantLabel(nom=nom))
    db.session.commit()
    return render_template('partials/ref_remettants_list.html',
                           **_ref_remettants_ctx(_current_role() == 'admin'))


@app.route('/api/referentiels/remettants/<int:item_id>/toggle', methods=['POST'])
@role_required('admin', 'saisie')
def api_ref_remettant_toggle(item_id):
    item = RemettantLabel.query.get_or_404(item_id)
    item.actif = not item.actif
    db.session.commit()
    return render_template('partials/ref_remettants_list.html',
                           **_ref_remettants_ctx(_current_role() == 'admin'))


@app.route('/api/referentiels/remettants/<int:item_id>/delete', methods=['DELETE', 'POST'])
@role_required('admin')
def api_ref_remettant_delete(item_id):
    item = RemettantLabel.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    return render_template('partials/ref_remettants_list.html',
                           **_ref_remettants_ctx(_current_role() == 'admin'))


# ─── LOGISTIQUE ──────────────────────────────────────────────────────────────

def _log_kpis():
    """Calcule les KPIs logistique (nombre d'entrées par statut)."""
    kpis = {'ARRIVÉ': 0, 'D10': 0, 'ÉCHÉANCE': 0, 'ARRIVE À ÉCHÉANCE': 0,
            'ÉCHU': 0, 'PAIEMENT EN COURS': 0, 'PAYÉ': 0, 'EN COURS': 0}
    for c in CommandeLogistique.query.all():
        s = c.statut
        kpis[s] = kpis.get(s, 0) + 1
    return kpis


def _get_logistique_notifications(limit=8):
    """Retourne les alertes d'échéance pour la gestion logistique."""
    today = date.today()
    alert_date = today + timedelta(days=7)

    # Workflow échéance : uniquement les entrées non payées.
    base_query = CommandeLogistique.query.filter(
        CommandeLogistique.date_echeance.isnot(None),
        CommandeLogistique.date_paiement.is_(None),
        CommandeLogistique.date_valeur.is_(None),
    )

    overdue_count = base_query.filter(CommandeLogistique.date_echeance < today).count()
    upcoming_count = base_query.filter(
        CommandeLogistique.date_echeance >= today,
        CommandeLogistique.date_echeance <= alert_date,
    ).count()

    critical_entries = base_query.filter(
        or_(
            CommandeLogistique.date_echeance < today,
            (
                (CommandeLogistique.date_echeance >= today) &
                (CommandeLogistique.date_echeance <= alert_date)
            ),
        )
    ).order_by(CommandeLogistique.date_echeance.asc(), CommandeLogistique.id.desc()).limit(limit).all()

    return {
        'today': today,
        'overdue_count': overdue_count,
        'upcoming_count': upcoming_count,
        'total_alerts': overdue_count + upcoming_count,
        'critical_entries': critical_entries,
    }


LOG_STATUTS = ['EN COURS', 'D10', 'ARRIVÉ', 'ÉCHÉANCE', 'ARRIVE À ÉCHÉANCE', 'ÉCHU', 'PAIEMENT EN COURS', 'PAYÉ']


@app.route('/api/logistique/bons')
@login_required
def api_logistique_bons_list():
    """HTMX partial: retourne la table des bons filtrée."""
    page     = request.args.get('page', 1, type=int)
    search   = request.args.get('search', '').strip()
    societe  = request.args.get('societe', '').strip()
    statut_f = request.args.get('statut', '').strip()

    q = BonCommande.query
    if search:
        q = q.filter(or_(
            BonCommande.fournisseur.ilike(f'%{search}%'),
            BonCommande.numero.ilike(f'%{search}%'),
        ))
    if societe:
        q = q.filter(BonCommande.societe == societe)
    if statut_f:
        q = q.filter(BonCommande.statut == statut_f)
    q = q.order_by(BonCommande.date_commande.desc(), BonCommande.id.desc())
    total       = q.count()
    bons        = q.offset((page - 1) * BON_PER_PAGE).limit(BON_PER_PAGE).all()
    total_pages = max(1, (total + BON_PER_PAGE - 1) // BON_PER_PAGE)
    return render_template('partials/logistique_bons_table.html',
                           bons=bons, page=page, total_pages=total_pages, total=total,
                           search=search, societe=societe, statut_f=statut_f,
                           bon_statuts=BON_STATUTS,
                           can_write=_current_role() in ('admin', 'saisie'),
                           is_admin=_current_role() == 'admin')


@app.route('/api/logistique/gestion')
@login_required
def api_logistique_gestion_list():
    """HTMX partial: retourne la table de gestion filtrée."""
    page     = request.args.get('page', 1, type=int)
    search   = request.args.get('search', '').strip()
    societe  = request.args.get('societe', '').strip()
    statut_f = request.args.get('statut', '').strip()
    date_filter = request.args.get('date_filter', '').strip()
    date_debut_raw = request.args.get('date_debut', '').strip()
    date_fin_raw = request.args.get('date_fin', '').strip()

    def _parse_date(v):
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except ValueError:
            return None

    date_debut = _parse_date(date_debut_raw)
    date_fin = _parse_date(date_fin_raw)
    date_fields = {
        'date_d10': CommandeLogistique.date_d10,
        'date_arrivee': CommandeLogistique.date_arrivee,
        'date_facture': CommandeLogistique.date_facture,
        'date_echeance': CommandeLogistique.date_echeance,
        'date_paiement': CommandeLogistique.date_paiement,
        'date_valeur': CommandeLogistique.date_valeur,
    }

    q = CommandeLogistique.query
    if search:
        q = q.filter(or_(
            CommandeLogistique.produit.ilike(f'%{search}%'),
            CommandeLogistique.fournisseur.ilike(f'%{search}%'),
            CommandeLogistique.remarque.ilike(f'%{search}%'),
        ))
    if societe:
        q = q.filter(CommandeLogistique.societe == societe)
    if date_filter in date_fields:
        df = date_fields[date_filter]
        if date_debut:
            q = q.filter(df >= date_debut)
        if date_fin:
            q = q.filter(df <= date_fin)
    q = q.order_by(CommandeLogistique.date_arrivee.desc().nullslast(), CommandeLogistique.id.desc())

    if statut_f:
        all_items  = q.all()
        filtered   = [c for c in all_items if c.statut == statut_f]
        total      = len(filtered)
        items      = filtered[(page - 1) * LOG_PER_PAGE: page * LOG_PER_PAGE]
    else:
        total = q.count()
        items = q.offset((page - 1) * LOG_PER_PAGE).limit(LOG_PER_PAGE).all()

    total_pages = max(1, (total + LOG_PER_PAGE - 1) // LOG_PER_PAGE)
    return render_template('partials/logistique_gestion_table.html',
                           items=items, page=page, total_pages=total_pages, total=total,
                           search=search, societe=societe,
                           statut_f=statut_f, date_filter=date_filter,
                           date_debut=date_debut_raw, date_fin=date_fin_raw,
                           today=date.today(),
                           can_write=_current_role() in ('admin', 'saisie'),
                           is_admin=_current_role() == 'admin')


@app.route('/logistique/gestion')
@login_required
def logistique_gestion():
    page     = request.args.get('page', 1, type=int)
    search   = request.args.get('search', '').strip()
    societe  = request.args.get('societe', '').strip()
    statut_f = request.args.get('statut', '').strip()
    date_filter = request.args.get('date_filter', '').strip()
    date_debut_raw = request.args.get('date_debut', '').strip()
    date_fin_raw = request.args.get('date_fin', '').strip()

    def _parse_date(v):
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except ValueError:
            return None

    date_debut = _parse_date(date_debut_raw)
    date_fin = _parse_date(date_fin_raw)
    date_fields = {
        'date_d10': CommandeLogistique.date_d10,
        'date_arrivee': CommandeLogistique.date_arrivee,
        'date_facture': CommandeLogistique.date_facture,
        'date_echeance': CommandeLogistique.date_echeance,
        'date_paiement': CommandeLogistique.date_paiement,
        'date_valeur': CommandeLogistique.date_valeur,
    }

    q = CommandeLogistique.query
    if search:
        q = q.filter(or_(
            CommandeLogistique.produit.ilike(f'%{search}%'),
            CommandeLogistique.fournisseur.ilike(f'%{search}%')
        ))
    if societe:
        q = q.filter(CommandeLogistique.societe == societe)
    if date_filter in date_fields:
        df = date_fields[date_filter]
        if date_debut:
            q = q.filter(df >= date_debut)
        if date_fin:
            q = q.filter(df <= date_fin)

    q = q.order_by(CommandeLogistique.date_arrivee.desc().nullslast(),
                   CommandeLogistique.id.desc())

    if statut_f:
        all_items        = q.all()
        items_filtered   = [c for c in all_items if c.statut == statut_f]
        total            = len(items_filtered)
        items            = items_filtered[(page - 1) * LOG_PER_PAGE: page * LOG_PER_PAGE]
    else:
        total = q.count()
        items = q.offset((page - 1) * LOG_PER_PAGE).limit(LOG_PER_PAGE).all()

    total_pages  = max(1, (total + LOG_PER_PAGE - 1) // LOG_PER_PAGE)

    return render_template('logistique_gestion.html',
                           items=items, page=page, total_pages=total_pages, total=total,
                           search=search, societe=societe,
                           statut_f=statut_f, date_filter=date_filter,
                           date_debut=date_debut_raw, date_fin=date_fin_raw,
                           kpis=_log_kpis(),
                           notifications=_get_logistique_notifications(),
                           today=date.today(),
                           log_statuts=LOG_STATUTS,
                           can_write=_current_role() in ('admin', 'saisie'),
                           is_admin=_current_role() == 'admin')


@app.route('/api/logistique/notifications')
@login_required
def api_logistique_notifications():
    return render_template(
        'partials/logistique_notifications_panel.html',
        notifications=_get_logistique_notifications(),
    )


def _log_form_fields(c):
    """Lit les champs du formulaire logistique depuis request.form et les applique à l'objet c."""
    def fd(k):
        v = request.form.get(k, '').strip()
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except ValueError:
            return None
    def ff(k):
        try:
            v = request.form.get(k, '').strip()
            return float(v) if v else None
        except ValueError:
            return None
    def fi(k):
        try:
            v = request.form.get(k, '').strip()
            return int(float(v)) if v else None
        except ValueError:
            return None

    if 'societe' in request.form:
        c.societe       = request.form.get('societe', '').strip()
    c.annee         = request.form.get('annee', '').strip() or None
    c.date_d10      = fd('date_d10')
    c.date_arrivee  = fd('date_arrivee')
    if 'fournisseur' in request.form:
        c.fournisseur   = request.form.get('fournisseur', '').strip().upper() or None
    c.produit       = request.form.get('produit', '').strip() or None
    c.emballage     = request.form.get('emballage', '').strip() or None
    c.quantite      = ff('quantite')
    c.tva           = ff('tva')
    c.montant_eur   = ff('montant_eur')
    c.cours         = ff('cours')
    c.date_facture  = fd('date_facture')
    c.code_paiement = request.form.get('code_paiement', '').strip() or None
    c.nb_jours      = fi('nb_jours')
    c.date_echeance = fd('date_echeance')
    c.date_paiement = fd('date_paiement')
    c.date_valeur   = fd('date_valeur')
    c.remarque      = request.form.get('remarque', '').strip() or None


@app.route('/api/logistique/add', methods=['POST'])
@role_required('admin', 'saisie')
def api_logistique_add():
    c = CommandeLogistique(cree_par=session.get('username', ''))
    _log_form_fields(c)
    db.session.add(c)
    db.session.commit()
    return redirect(url_for('logistique_gestion'))


@app.route('/api/logistique/<int:item_id>/edit', methods=['POST'])
@role_required('admin', 'saisie')
def api_logistique_edit(item_id):
    c = CommandeLogistique.query.get_or_404(item_id)
    _log_form_fields(c)
    db.session.commit()
    return redirect(url_for('logistique_gestion'))


@app.route('/api/logistique/<int:item_id>/delete', methods=['POST', 'DELETE'])
@role_required('admin')
def api_logistique_delete(item_id):
    c = CommandeLogistique.query.get_or_404(item_id)
    db.session.delete(c)
    db.session.commit()
    return redirect(url_for('logistique_gestion'))


# ── Products (Referentiels) ──────────────────────────────────────────────────

@app.route('/api/products/by-company')
@login_required
def api_products_by_company():
    """Get products filtered by company for autofill."""
    company = request.args.get('company', '').strip()
    if not company:
        return jsonify([])
    
    products = Product.query.filter_by(company=company).order_by(Product.reference).all()
    return jsonify([p.to_dict() for p in products])


@app.route('/api/products/search')
@login_required
def api_products_search():
    """Search products by reference or designation."""
    q = request.args.get('q', '').strip()
    company = request.args.get('company', '').strip()
    
    query = Product.query
    if company:
        query = query.filter_by(company=company)
    
    if q:
        query = query.filter(or_(
            Product.reference.ilike(f'%{q}%'),
            Product.designation.ilike(f'%{q}%')
        ))
    
    products = query.order_by(Product.reference).limit(50).all()
    return jsonify([p.to_dict() for p in products])


# ── Bons de commande ──────────────────────────────────────────────────────────

@app.route('/logistique/bons')
@login_required
def logistique_bons():
    page     = request.args.get('page', 1, type=int)
    search   = request.args.get('search', '').strip()
    societe  = request.args.get('societe', '').strip()
    statut_f = request.args.get('statut', '').strip()

    q = BonCommande.query
    if search:
        q = q.filter(or_(
            BonCommande.fournisseur.ilike(f'%{search}%'),
            BonCommande.numero.ilike(f'%{search}%'),
            BonCommande.notes.ilike(f'%{search}%')
        ))
    if societe:
        q = q.filter(BonCommande.societe == societe)
    if statut_f:
        q = q.filter(BonCommande.statut == statut_f)

    q = q.order_by(BonCommande.date_commande.desc(), BonCommande.id.desc())
    total       = q.count()
    bons        = q.offset((page - 1) * BON_PER_PAGE).limit(BON_PER_PAGE).all()
    total_pages = max(1, (total + BON_PER_PAGE - 1) // BON_PER_PAGE)

    fournisseurs = [r[0] for r in db.session.query(CommandeLogistique.fournisseur)
                    .filter(CommandeLogistique.fournisseur.isnot(None), CommandeLogistique.fournisseur != '')
                    .distinct().order_by(CommandeLogistique.fournisseur).all()]

    import json as _json
    products = Product.query.order_by(Product.company, Product.designation).all()
    products_json = _json.dumps([p.to_dict() for p in products])

    return render_template('logistique_bons.html',
                           bons=bons, page=page, total_pages=total_pages, total=total,
                           search=search, societe=societe, statut_f=statut_f,
                           bon_statuts=BON_STATUTS, fournisseurs=fournisseurs,
                           products_json=products_json,
                           can_write=_current_role() in ('admin', 'saisie'),
                           is_admin=_current_role() == 'admin')


@app.route('/api/logistique/bons/add', methods=['POST'])
@role_required('admin', 'saisie')
def api_bon_add():
    def fd(k):
        v = request.form.get(k, '').strip()
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except ValueError:
            return None

    year_str = str(date.today().year)
    count    = BonCommande.query.filter(BonCommande.numero.like(f'BC-{year_str}-%')).count()
    numero   = f'BC-{year_str}-{count + 1:04d}'

    bon = BonCommande(
        numero                = numero,
        societe               = request.form.get('societe', '').strip(),
        fournisseur           = request.form.get('fournisseur', '').strip().upper() or None,
        statut                = request.form.get('statut', 'Brouillon'),
        date_commande         = fd('date_commande') or date.today(),
        date_livraison_prevue = fd('date_livraison_prevue'),
        notes                 = request.form.get('notes', '').strip() or None,
        cree_par              = session.get('username', ''),
    )
    db.session.add(bon)
    db.session.flush()

    designations   = request.form.getlist('designation[]')
    quantites      = request.form.getlist('quantite[]')
    unites         = request.form.getlist('unite[]')
    prix_unitaires = request.form.getlist('prix_unitaire[]')
    references     = request.form.getlist('reference[]')

    for i, desig in enumerate(designations):
        desig = desig.strip()
        if not desig:
            continue
        try:
            qty = float(quantites[i]) if i < len(quantites) and quantites[i].strip() else 1.0
        except (ValueError, IndexError):
            qty = 1.0
        try:
            prix = float(prix_unitaires[i]) if i < len(prix_unitaires) and prix_unitaires[i].strip() else None
        except (ValueError, IndexError):
            prix = None
        ref   = references[i].strip() if i < len(references) else ''
        unite = unites[i].strip() if i < len(unites) else ''
        db.session.add(LigneCommande(bon_id=bon.id, reference=ref or None,
                                     designation=desig, quantite=qty,
                                     unite=unite or None, prix_unitaire=prix))
    db.session.commit()

    # Créer automatiquement l'entrée dans CommandeLogistique
    total_montant = sum(
        (float(quantites[i]) if i < len(quantites) and quantites[i].strip() else 1.0) *
        (float(prix_unitaires[i]) if i < len(prix_unitaires) and prix_unitaires[i].strip() else 0)
        for i in range(len(designations))
        if designations[i].strip()
    )
    
    log_entry = CommandeLogistique(
        bon_id            = bon.id,
        ref_log           = numero,  # Même numéro que le bon
        societe           = bon.societe,
        annee             = str(date.today().year),
        fournisseur       = bon.fournisseur,
        montant_eur       = total_montant if total_montant else None,
        cree_par          = session.get('username', ''),
    )
    db.session.add(log_entry)
    db.session.commit()

    return redirect(url_for('logistique_bons'))


@app.route('/api/logistique/bons/<int:bon_id>/statut', methods=['POST'])
@role_required('admin', 'saisie')
def api_bon_statut(bon_id):
    bon = BonCommande.query.get_or_404(bon_id)
    bon.statut = request.form.get('statut', bon.statut)
    db.session.commit()
    return redirect(url_for('logistique_bons'))


@app.route('/api/logistique/bons/<int:bon_id>/update', methods=['POST'])
@role_required('admin', 'saisie')
def api_bon_update(bon_id):
    bon = BonCommande.query.get_or_404(bon_id)

    def fd(k):
        v = request.form.get(k, '').strip()
        try:
            return datetime.strptime(v, '%Y-%m-%d').date() if v else None
        except ValueError:
            return None

    bon.societe       = request.form.get('societe', bon.societe).strip()
    bon.fournisseur   = request.form.get('fournisseur', '').strip().upper() or None
    bon.statut        = request.form.get('statut', bon.statut)
    bon.date_commande = fd('date_commande') or bon.date_commande
    bon.notes         = request.form.get('notes', '').strip() or None

    # Remplacer les lignes existantes
    LigneCommande.query.filter_by(bon_id=bon.id).delete()
    db.session.flush()

    designations   = request.form.getlist('designation[]')
    quantites      = request.form.getlist('quantite[]')
    prix_unitaires = request.form.getlist('prix_unitaire[]')
    references     = request.form.getlist('reference[]')

    total_montant = 0.0
    for i, desig in enumerate(designations):
        desig = desig.strip()
        if not desig:
            continue
        try:
            qty = float(quantites[i]) if i < len(quantites) and quantites[i].strip() else 1.0
        except (ValueError, IndexError):
            qty = 1.0
        try:
            prix = float(prix_unitaires[i]) if i < len(prix_unitaires) and prix_unitaires[i].strip() else None
        except (ValueError, IndexError):
            prix = None
        ref = references[i].strip() if i < len(references) else ''
        db.session.add(LigneCommande(bon_id=bon.id, reference=ref or None,
                                     designation=desig, quantite=qty,
                                     prix_unitaire=prix))
        total_montant += qty * (prix or 0)

    # Mettre à jour l'entrée CommandeLogistique associée
    log_entry = CommandeLogistique.query.filter_by(bon_id=bon.id).first()
    if log_entry:
        log_entry.societe     = bon.societe
        log_entry.fournisseur = bon.fournisseur
        log_entry.montant_eur = total_montant or None

    db.session.commit()
    return redirect(url_for('logistique_bons'))


@app.route('/api/logistique/bons/<int:bon_id>/delete', methods=['POST', 'DELETE'])
@role_required('admin')
def api_bon_delete(bon_id):
    bon = BonCommande.query.get_or_404(bon_id)
    db.session.delete(bon)
    db.session.commit()
    return redirect(url_for('logistique_bons'))


@app.route('/logistique/bons/<int:bon_id>/print')
@login_required
def print_bon(bon_id):
    """Print bon de commande."""
    bon = BonCommande.query.get_or_404(bon_id)
    from datetime import datetime
    return render_template('bon_print.html', bon=bon, now=datetime.now())


@app.route('/api/logistique/bons/<int:bon_id>/detail')
@login_required
def api_bon_detail(bon_id):
    """Return bon de commande data as JSON (used by gestion page viewer)."""
    bon = BonCommande.query.get_or_404(bon_id)
    return jsonify(bon.to_dict())


# --- Suppression ---

@app.route('/delete/<int:op_id>', methods=['DELETE', 'POST'])
@role_required('admin')
def delete_operation(op_id):
    op = Operation.query.get_or_404(op_id)
    _log_audit(op.id, 'suppression', f"{op.type_operation} - {op.client} - {op.montant}")
    db.session.delete(op)
    db.session.commit()
    if request.headers.get('HX-Request'):
        return '<div class="alert alert-info fade-in"><i class="fas fa-trash mr-2"></i>Opération supprimée.</div>'
    flash('Opération supprimée.', 'info')
    return redirect(url_for('consultation'))


# --- Consultation ---

def _build_operations_query(args):
    """Construit et retourne la query + pagination à partir d'un dict de params."""
    query = Operation.query

    search = args.get('search', '').strip()
    if search:
        query = query.filter(or_(
            Operation.client.ilike(f'%{search}%'),
            Operation.remettant.ilike(f'%{search}%'),
            Operation.banque.ilike(f'%{search}%'),
            Operation.numero_piece.ilike(f'%{search}%'),
            Operation.remarque.ilike(f'%{search}%'),
            Operation.societe.ilike(f'%{search}%'),
        ))

    type_op = args.get('type_operation', '').strip()
    if type_op:
        query = query.filter(Operation.type_operation == type_op)

    type_cheque = args.get('type_cheque', '').strip()
    if type_cheque:
        query = query.filter(Operation.type_operation == 'Chèque', Operation.type_detail == type_cheque)

    societe = args.get('societe', '').strip()
    if societe:
        query = query.filter(Operation.societe == societe)

    statut = args.get('statut', '').strip()
    if statut:
        query = query.filter(Operation.statut == statut)

    date_filter = args.get('date_filter', 'date_operation').strip()
    date_column = Operation.date_operation
    if date_filter == 'date_reception':
        query = query.filter(Operation.type_operation == 'Chèque')
        date_column = Operation.date_reception
    elif date_filter == 'date_echeance':
        query = query.filter(
            Operation.type_operation == 'Chèque',
            Operation.type_detail == 'À échéance',
        )
        date_column = Operation.date_encaissement

    date_debut = args.get('date_debut', '').strip()
    if date_debut:
        query = query.filter(date_column >= date_debut)

    date_fin = args.get('date_fin', '').strip()
    if date_fin:
        query = query.filter(date_column <= date_fin)

    # Tri
    sort = args.get('sort', '').strip()
    order = args.get('order', 'desc')
    if sort and hasattr(Operation, sort):
        col = getattr(Operation, sort)
        query = query.order_by(col.desc() if order == 'desc' else col.asc())
    else:
        today = date.today()
        alert_date = today + timedelta(days=7)
        urgency_rank = case(
            (
                (
                    (Operation.type_operation == 'Chèque') &
                    (Operation.type_detail == 'À échéance') &
                    (Operation.date_encaissement.isnot(None)) &
                    (Operation.date_encaissement < today)
                ),
                0,
            ),
            (
                (
                    (Operation.type_operation == 'Chèque') &
                    (Operation.type_detail == 'À échéance') &
                    (Operation.date_encaissement.isnot(None)) &
                    (Operation.date_encaissement >= today) &
                    (Operation.date_encaissement <= alert_date)
                ),
                1,
            ),
            else_=2,
        )
        query = query.order_by(
            urgency_rank.asc(),
            Operation.date_encaissement.asc(),
            Operation.date_operation.desc(),
        )

    total_montant = db.session.query(func.sum(Operation.montant)).filter(
        Operation.id.in_(query.with_entities(Operation.id))
    ).scalar() or 0
    total_count = query.count()

    page = int(args.get('page', 1) or 1)
    per_page = 25
    offset = (page - 1) * per_page
    operations = query.limit(per_page).offset(offset).all()
    total_pages = (total_count + per_page - 1) // per_page

    return {
        'operations': operations,
        'total_montant': total_montant,
        'total_count': total_count,
        'page': page,
        'total_pages': total_pages,
    }


@app.route('/consultation')
@login_required
def consultation():
    _auto_update_echeance_statuts()
    notifications = _get_echeance_notifications()
    role = _current_role()
    rejections = _get_recent_rejections() if role != 'admin' else {'rejections': [], 'total': 0}
    ops_data = _build_operations_query(request.args)
    return render_template(
        'consultation.html',
        notifications=notifications,
        rejections=rejections,
        is_admin=(role == 'admin'),
        **ops_data,
    )


@app.route('/api/operations')
@login_required
def api_operations():
    data = _build_operations_query(request.args)
    return render_template(
        'partials/operations_table.html',
        is_admin=_current_role() == 'admin',
        status_choices=STATUS_CHOICES,
        **data,
    )


@app.route('/api/notifications')
@login_required
def api_notifications():
    notifications = _get_echeance_notifications()
    return render_template(
        'partials/notifications_panel.html',
        notifications=notifications,
        is_admin=_current_role() == 'admin',
    )


@app.route('/api/notifications/badge')
@login_required
def api_notifications_badge():
    notifications = _get_echeance_notifications()
    log_notifications = _get_logistique_notifications()
    # Pour les non-admins, ajouter le compte des rejets récents
    rejections_count = 0 if _current_role() == 'admin' else _get_recent_rejections()['total']
    consultation_total = (notifications.get('total_alerts', 0) if notifications else 0) + rejections_count
    logistique_total = log_notifications.get('total_alerts', 0) if log_notifications else 0
    target_url = url_for('consultation') if consultation_total > 0 else (
        url_for('logistique_gestion') if logistique_total > 0 else url_for('consultation')
    )
    return render_template(
        'partials/global_notifications_badge.html',
        notifications=notifications,
        log_notifications=log_notifications,
        rejections_count=rejections_count,
        target_url=target_url,
    )


@app.route('/api/rejections')
@login_required
def api_rejections():
    """Retourne les rejets récents (sauf pour les admins)."""
    if _current_role() == 'admin':

        rejections_data = {'rejections': [], 'total': 0}  # Admins ne voient pas les rejets
    else:
        rejections_data = _get_recent_rejections()
    
    return render_template(
        'partials/rejections_panel.html',
        rejections=rejections_data,
    )

@app.route('/api/operations/<int:op_id>')
@login_required
def api_operation_detail(op_id):
    op = Operation.query.get_or_404(op_id)
    audits = AuditLog.query.filter_by(operation_id=op_id).order_by(AuditLog.date_action.desc()).all()
    if request.headers.get('HX-Request'):
        return render_template('partials/operation_detail.html', operation=op, audits=audits)
    return jsonify(op.to_dict())


@app.route('/api/operations/<int:op_id>/status', methods=['POST'])
@role_required('admin')
def api_update_operation_status(op_id):
    op = Operation.query.get_or_404(op_id)
    new_status = (request.form.get('statut') or '').strip()

    if new_status not in STATUS_CHOICES:
        return ('Statut invalide.', 400)

    if op.statut == new_status:
        return ('', 204)

    old_status = op.statut
    op.statut = new_status
    op.date_modification = datetime.utcnow()
    db.session.commit()

    # Log audit avec message spécial si c'est un rejet
    if new_status == 'Rejeté':
        audit_message = f"Statut modifié: {old_status} -> Rejeté (Chèque #{op.numero_piece} de {op.client} rejeté)"
    else:
        audit_message = f'Statut modifié: {old_status} -> {new_status}'
    
    _log_audit(op.id, 'modification', audit_message)
    return ('', 204)


# --- Import Excel ---

@app.route('/import', methods=['GET', 'POST'])
@role_required('admin', 'saisie')
def import_excel():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename.endswith(('.xlsx', '.xls')):
            flash('Veuillez sélectionner un fichier Excel (.xlsx)', 'error')
            return redirect(url_for('import_excel'))

        try:
            wb = load_workbook(file, read_only=True, data_only=True)
            imported = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                for row in rows:
                    if not row or all(cell is None for cell in row):
                        continue
                    op = _parse_excel_row(row, sheet_name)
                    if op:
                        db.session.add(op)
                        imported += 1
            db.session.commit()
            flash(f'{imported} opérations importées avec succès !', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de l\'import : {str(e)}', 'error')

        return redirect(url_for('import_excel'))

    return render_template('import.html')


# --- Historique ---

@app.route('/historique')
@login_required
def historique():
    audits = AuditLog.query.order_by(AuditLog.date_action.desc()).limit(200).all()
    return render_template('historique.html', audits=audits)


# --- Export ---

@app.route('/export')
@login_required
def export_excel():
    query = Operation.query
    search = request.args.get('search', '').strip()
    if search:
        query = query.filter(or_(
            Operation.client.ilike(f'%{search}%'),
            Operation.remettant.ilike(f'%{search}%'),
            Operation.banque.ilike(f'%{search}%'),
            Operation.numero_piece.ilike(f'%{search}%'),
            Operation.remarque.ilike(f'%{search}%'),
            Operation.societe.ilike(f'%{search}%'),
        ))

    type_op = request.args.get('type_operation', '').strip()
    if type_op:
        query = query.filter(Operation.type_operation == type_op)

    type_cheque = request.args.get('type_cheque', '').strip()
    if type_cheque:
        query = query.filter(Operation.type_operation == 'Chèque', Operation.type_detail == type_cheque)

    societe = request.args.get('societe', '').strip()
    if societe:
        query = query.filter(Operation.societe == societe)

    statut = request.args.get('statut', '').strip()
    if statut:
        query = query.filter(Operation.statut == statut)

    date_debut = request.args.get('date_debut', '').strip()
    if date_debut:
        query = query.filter(Operation.date_operation >= date_debut)

    date_fin = request.args.get('date_fin', '').strip()
    if date_fin:
        query = query.filter(Operation.date_operation <= date_fin)

    date_echeance = request.args.get('date_echeance', '').strip()
    if date_echeance:
        query = query.filter(
            Operation.type_operation == 'Chèque',
            Operation.type_detail == 'À échéance',
            Operation.date_encaissement == date_echeance,
        )

    operations = query.order_by(Operation.date_operation.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Opérations"
    headers = ['ID', 'Date', 'Type', 'Société', 'Client', 'Remettant', 'Montant',
               'Banque', 'N° Pièce', 'Statut', 'Famille', 'Remarque', 'Saisi par']
    ws.append(headers)
    for op in operations:
        ws.append([
            op.id,
            op.date_operation.strftime('%d/%m/%Y') if op.date_operation else '',
            op.type_operation, op.societe, op.client, op.remettant or '',
            op.montant, op.banque or '', op.numero_piece or '',
            op.statut, op.famille or '', op.remarque or '', op.cree_par or '',
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'operations_{date.today().strftime("%Y%m%d")}.xlsx'
    file_bytes = output.getvalue()
    response = make_response(file_bytes)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.headers['Content-Length'] = str(len(file_bytes))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response


# --- Utilitaires ---

def _parse_date(date_str):
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _log_audit(operation_id, action, details=''):
    log = AuditLog(
        operation_id=operation_id,
        action=action,
        utilisateur=session.get('user_nom', 'Système'),
        details=details,
    )
    db.session.add(log)
    db.session.commit()


def _parse_excel_row(row, sheet_name):
    """Parse une ligne Excel en opération"""
    try:
        sheet_lower = sheet_name.lower()
        if 'chèque' in sheet_lower or 'cheque' in sheet_lower:
            type_op = 'Chèque'
        elif 'virement' in sheet_lower:
            type_op = 'Virement'
        elif 'versement' in sheet_lower:
            type_op = 'Versement'
        else:
            type_op = 'Versement'

        date_val = None
        montant_val = None
        client_val = None
        banque_val = None

        for cell in row:
            if isinstance(cell, datetime):
                date_val = cell.date()
            elif isinstance(cell, date):
                date_val = cell
            elif isinstance(cell, (int, float)) and cell > 10 and montant_val is None:
                montant_val = float(cell)
            elif isinstance(cell, str) and len(cell) > 2:
                if client_val is None:
                    client_val = cell
                elif banque_val is None:
                    banque_val = cell

        if not date_val or not montant_val or not client_val:
            return None

        return Operation(
            type_operation=type_op, societe='ENT',
            date_operation=date_val, client=client_val,
            montant=montant_val, banque=_normalize_bank_name(banque_val),
            statut='Encaissé', cree_par='Import Excel',
        )
    except Exception:
        return None


# --- Initialisation ---
with app.app_context():
    db.create_all()

    # Normalise legacy roles to the new 3-role model.
    for u in User.query.all():
        normalized = _normalize_role(u.role)
        if u.role != normalized:
            u.role = normalized

    # Rename legacy admin account boss -> mehdi (keeping password hash).
    legacy_boss = User.query.filter_by(username='boss').first()
    mehdi_user = User.query.filter_by(username='mehdi').first()
    if legacy_boss and not mehdi_user:
        legacy_boss.username = 'mehdi'
        if not legacy_boss.nom_complet or legacy_boss.nom_complet.strip().lower() == 'boss':
            legacy_boss.nom_complet = 'Mehdi'
        legacy_boss.role = 'admin'

    if not User.query.filter_by(username='mehdi').first():
        mehdi = User(username='mehdi', nom_complet='Mehdi', role='admin')
        mehdi.set_password('srid2024mehdi')
        db.session.add(mehdi)

    if not User.query.filter_by(username='sabrina').first():
        sabrina = User(username='sabrina', nom_complet='Sabrina', role='saisie')
        sabrina.set_password('srid2024sab')
        db.session.add(sabrina)
    else:
        sabrina = User.query.filter_by(username='sabrina').first()
        sabrina.role = 'saisie'

    # Normalize legacy statuses to the approved status list.
    for op in Operation.query.all():
        normalized_statut = _normalize_legacy_statut(op)
        if op.statut != normalized_statut:
            op.statut = normalized_statut

    # Seed référentiels depuis les opérations existantes (idempotent).
    existing_clients = {c.nom.lower() for c in ClientLabel.query.all()}
    client_vals = db.session.query(Operation.client).filter(
        Operation.client.isnot(None), Operation.client != ''
    ).distinct().all()
    for (nom,) in client_vals:
        nom = (nom or '').strip()
        if nom and nom.lower() not in existing_clients:
            db.session.add(ClientLabel(nom=nom))
            existing_clients.add(nom.lower())

    existing_remettants = {r.nom.lower() for r in RemettantLabel.query.all()}
    remettant_vals = db.session.query(Operation.remettant).filter(
        Operation.remettant.isnot(None), Operation.remettant != ''
    ).distinct().all()
    for (nom,) in remettant_vals:
        nom = (nom or '').strip()
        if nom and nom.lower() not in existing_remettants:
            db.session.add(RemettantLabel(nom=nom))
            existing_remettants.add(nom.lower())

    db.session.commit()

    # Seed logistique depuis bon-md/logistics-data.js (idempotent, une seule fois)
    if CommandeLogistique.query.count() == 0:
        _js_path = os.path.abspath(os.path.join(
            os.path.dirname(os.path.abspath(__file__)), '..', 'bon-md', 'logistics-data.js'))
        if os.path.exists(_js_path):
            with open(_js_path, 'r', encoding='utf-8') as _f:
                _content = _f.read()
            _m = re.search(r'window\.LOGISTICS_SEED\s*=\s*(\[.*?\]);', _content, re.DOTALL)
            if _m:
                _data = json.loads(_m.group(1))
                def _pd(s):
                    if not s or not isinstance(s, str) or not s.strip():
                        return None
                    try:
                        return datetime.strptime(s.strip(), '%Y-%m-%d').date()
                    except ValueError:
                        return None
                def _pf(v):
                    try:
                        return float(v) if v is not None and v != '' else None
                    except (TypeError, ValueError):
                        return None
                def _pi(v):
                    try:
                        return int(float(v)) if v is not None and v != '' else None
                    except (TypeError, ValueError):
                        return None
                for _item in _data:
                    db.session.add(CommandeLogistique(
                        ref_log       = _item.get('id'),
                        societe       = _item.get('company', ''),
                        annee         = str(_item.get('year', '')) if _item.get('year') else None,
                        date_d10      = _pd(_item.get('dateD10')),
                        date_arrivee  = _pd(_item.get('arrivalDate')),
                        fournisseur   = (_item.get('supplier') or '').upper() or None,
                        produit       = _item.get('product') or None,
                        emballage     = _item.get('packaging') or None,
                        quantite      = _pf(_item.get('quantity')),
                        tva           = _pf(_item.get('vat')),
                        montant_eur   = _pf(_item.get('amountEur')),
                        cours         = _pf(_item.get('rate')),
                        date_facture  = _pd(_item.get('invoiceDate')),
                        code_paiement = _item.get('paymentCode') or None,
                        nb_jours      = _pi(_item.get('paymentDays')),
                        date_echeance = _pd(_item.get('dueDate')),
                        date_paiement = _pd(_item.get('paymentDate')),
                        date_valeur   = _pd(_item.get('valueDate')),
                        remarque      = _item.get('remark') or None,
                        cree_par      = 'seed',
                    ))
                db.session.commit()


if __name__ == '__main__':
    app.run(debug=True, port=5000)
